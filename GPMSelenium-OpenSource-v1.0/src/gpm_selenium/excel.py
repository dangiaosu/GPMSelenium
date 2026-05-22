from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any, Mapping

import openpyxl

STATUS_COLUMN_NAME: str = "Status"
LEGACY_SUCCESS_STATUS: str = "OKIE"
SUCCESS_STATUS: str = "SUCCESS"
SUCCESS_STATUSES: set[str] = {LEGACY_SUCCESS_STATUS, SUCCESS_STATUS}


class ExcelDataError(ValueError):
    pass


@dataclass(frozen=True)
class ExcelRow:
    row_number: int
    values: dict[str, str]
    status: str


@dataclass(frozen=True)
class ExcelPreview:
    total_rows: int
    pending_rows: int
    skipped_rows: int
    error_rows: int
    missing_columns: list[str]


def read_excel_rows(excel_path: Path, required_columns: list[str]) -> list[ExcelRow]:
    rows, headers = read_sheet_values(excel_path)
    if len(rows) == 0:
        raise ExcelDataError(f"Excel file is empty; excel_path={excel_path}")

    column_indexes: dict[str, int] = build_column_indexes(headers)
    missing_columns: list[str] = [column_name for column_name in required_columns if column_name not in column_indexes]
    if len(missing_columns) > 0:
        raise ExcelDataError(f"Excel file missing required columns; excel_path={excel_path}; columns={missing_columns}")

    status_index: int | None = column_indexes.get(STATUS_COLUMN_NAME)
    result: list[ExcelRow] = []
    for zero_based_index, raw_row in enumerate(rows[1:], start=1):
        row_number: int = zero_based_index + 1
        if row_is_empty(raw_row):
            continue
        values: dict[str, str] = {
            header: read_cell(raw_row, index) for header, index in column_indexes.items() if header != STATUS_COLUMN_NAME
        }
        status: str = read_cell(raw_row, status_index) if status_index is not None else ""
        result.append(ExcelRow(row_number=row_number, values=values, status=status))
    return result


def preview_excel(excel_path: Path, required_columns: list[str]) -> ExcelPreview:
    rows, headers = read_sheet_values(excel_path)
    column_indexes: dict[str, int] = build_column_indexes(headers) if len(headers) > 0 else {}
    missing_columns: list[str] = [column_name for column_name in required_columns if column_name not in column_indexes]
    status_index: int | None = column_indexes.get(STATUS_COLUMN_NAME)
    total_rows: int = 0
    skipped_rows: int = 0
    error_rows: int = 0
    pending_rows: int = 0
    for raw_row in rows[1:]:
        if row_is_empty(raw_row):
            continue
        total_rows += 1
        status: str = read_cell(raw_row, status_index) if status_index is not None else ""
        if is_success_status(status):
            skipped_rows += 1
        else:
            pending_rows += 1
            if status != "":
                error_rows += 1
    return ExcelPreview(
        total_rows=total_rows,
        pending_rows=pending_rows,
        skipped_rows=skipped_rows,
        error_rows=error_rows,
        missing_columns=missing_columns,
    )


def pending_rows(excel_rows: list[ExcelRow]) -> list[ExcelRow]:
    return [row for row in excel_rows if not is_success_status(row.status)]


def is_success_status(status: str) -> bool:
    return status.strip().upper() in SUCCESS_STATUSES


def initialize_status_column(excel_path: Path) -> int:
    workbook: openpyxl.Workbook = openpyxl.load_workbook(excel_path)
    try:
        worksheet: openpyxl.worksheet.worksheet.Worksheet = workbook.active
        status_column: int = find_or_create_status_column(worksheet)
        workbook.save(excel_path)
        return status_column
    finally:
        workbook.close()


def write_status(excel_path: Path, row_number: int, status: str) -> None:
    workbook: openpyxl.Workbook = openpyxl.load_workbook(excel_path)
    try:
        worksheet: openpyxl.worksheet.worksheet.Worksheet = workbook.active
        status_column: int = find_or_create_status_column(worksheet)
        worksheet.cell(row=row_number, column=status_column, value=status)
        workbook.save(excel_path)
    finally:
        workbook.close()


def read_sheet_values(excel_path: Path) -> tuple[list[tuple[Any, ...]], tuple[Any, ...]]:
    if not excel_path.exists():
        raise ExcelDataError(f"Excel file does not exist; excel_path={excel_path}")
    workbook: openpyxl.Workbook = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    try:
        worksheet: openpyxl.worksheet.worksheet.Worksheet = workbook.active
        rows: list[tuple[Any, ...]] = list(worksheet.iter_rows(values_only=True))
    finally:
        workbook.close()
    headers: tuple[Any, ...] = rows[0] if len(rows) > 0 else tuple()
    return rows, headers


def build_column_indexes(header_values: tuple[Any, ...]) -> dict[str, int]:
    indexes: dict[str, int] = {}
    for column_index, raw_header in enumerate(header_values):
        if raw_header is None:
            continue
        header: str = str(raw_header).strip()
        if header != "":
            indexes[header] = column_index
    return indexes


def row_is_empty(row_values: tuple[Any, ...]) -> bool:
    return all(raw_value is None or str(raw_value).strip() == "" for raw_value in row_values)


def read_cell(row_values: tuple[Any, ...], column_index: int | None) -> str:
    if column_index is None or column_index >= len(row_values):
        return ""
    raw_value: Any = row_values[column_index]
    if raw_value is None:
        return ""
    if isinstance(raw_value, int):
        return str(raw_value)
    if isinstance(raw_value, float) and raw_value.is_integer():
        return str(int(raw_value))
    return str(raw_value).strip()


def find_or_create_status_column(worksheet: openpyxl.worksheet.worksheet.Worksheet) -> int:
    for column_index in range(1, worksheet.max_column + 1):
        raw_header: Any = worksheet.cell(row=1, column=column_index).value
        if raw_header is not None and str(raw_header).strip().lower() == STATUS_COLUMN_NAME.lower():
            return column_index
    column_index = worksheet.max_column + 1
    worksheet.cell(row=1, column=column_index, value=STATUS_COLUMN_NAME)
    return column_index


def row_profile_id(row: ExcelRow) -> str:
    return require_value(row.values, "ProfileID", row.row_number)


def row_profile_name(row: ExcelRow) -> str:
    return row.values.get("ProfileName", f"Row {row.row_number}")


def require_value(values: Mapping[str, str], column_name: str, row_number: int) -> str:
    value: str = values.get(column_name, "").strip()
    if value == "":
        raise ExcelDataError(f"Excel cell is required; row_number={row_number}; column_name={column_name}")
    return value
